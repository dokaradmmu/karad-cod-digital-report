"""
Karad Division — COD Digital Transaction % Report
All Excel generation logic lives here. app.py only handles the Streamlit UI.
"""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Constants ────────────────────────────────────────────────────────────────
NAVY = "1F3864"
LIGHT_BLUE = "DCE6F1"
WHITE = "FFFFFF"
RED = "C00000"
GOLD = "BF8F00"
DARK_GREEN = "375623"
DARK_ORANGE = "BF6000"

THIN = Side(style="thin", color="AAAAAA")
ALL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

OFFICE_ROW_SIZE = 12   # office/data rows — deliberately larger than headers
HEADER_SIZE = 10
TITLE_SIZE = 18
SECTION_SIZE = 12

TYPE_DISPLAY = {"BPO": "B.O", "SPO": "S.O", "HPO": "H.O"}
TYPE_SORT_RANK = {"BPO": 0, "SPO": 1, "HPO": 1}


# ── Style helpers ────────────────────────────────────────────────────────────
def _font(bold=False, color="000000", size=OFFICE_ROW_SIZE, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _align(h="center", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


def _merge_write(ws, row, col_start, col_end, value, font, fill, align, height=None):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start)
    cell.value = value
    cell.font = font
    cell.fill = fill
    cell.alignment = align
    cell.border = ALL_BORDER
    for c in range(col_start + 1, col_end + 1):
        ws.cell(row=row, column=c).border = ALL_BORDER
    if height:
        ws.row_dimensions[row].height = height


def _write_header_row(ws, row, headers, widths, height=40, size=HEADER_SIZE):
    for c, (hdr, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=c, value=hdr)
        cell.font = _font(bold=True, color="FFFFFF", size=size)
        cell.fill = _fill(NAVY)
        cell.alignment = _align("center")
        cell.border = ALL_BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[row].height = height


# ── Loaders ──────────────────────────────────────────────────────────────────
def load_master(file_obj):
    df = pd.read_excel(file_obj, dtype=str)
    df.columns = df.columns.str.strip()
    for col in ["Sub Division Name", "Sub Office Name", "Office ID", "Office Name", "Office Type Code"]:
        df[col] = df[col].str.strip()
    df["Office ID"] = pd.to_numeric(df["Office ID"], errors="coerce")
    return df[["Sub Division Name", "Sub Office Name", "Office ID", "Office Name", "Office Type Code"]]


def load_cod_csv(file_obj):
    df = pd.read_csv(file_obj, dtype=str)
    df.columns = df.columns.str.strip()
    df["Office ID"] = pd.to_numeric(df["Office ID"], errors="coerce")
    for col in ["Total COD Count", "COD Digital Count"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df.groupby("Office ID", as_index=False).agg(
        Total=("Total COD Count", "sum"),
        Digital=("COD Digital Count", "sum"),
    )


# ── Data-quality auto-correction ────────────────────────────────────────────
def reconcile_sub_division(master: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """
    For each Sub Office Name, check whether the parent S.O./H.O. row and all its
    B.O. rows agree on Sub Division Name. If not, force the majority value among
    the B.O. rows onto the whole group and log the correction.

    Also detects genuine cross-division name collisions (same Sub Office Name,
    truly two different sub-offices) — these are reported but NOT merged, since
    grouping elsewhere always uses (Sub Division Name + Sub Office Name) together.
    """
    df = master.copy()
    corrections = []

    for so_name, grp in df.groupby("Sub Office Name"):
        divisions = grp["Sub Division Name"].unique()
        if len(divisions) <= 1:
            continue

        bo_rows = grp[grp["Office Type Code"] == "BPO"]
        if len(bo_rows) > 0:
            majority_div = bo_rows["Sub Division Name"].value_counts().idxmax()
        else:
            majority_div = grp["Sub Division Name"].value_counts().idxmax()

        # Only auto-correct if it looks like a tagging error: i.e. the B.O.s
        # themselves are unanimous, but the S.O./H.O. row alone disagrees.
        bo_divisions = bo_rows["Sub Division Name"].unique() if len(bo_rows) > 0 else []
        if len(bo_divisions) == 1:
            mismatched = grp[grp["Sub Division Name"] != majority_div]
            for _, row in mismatched.iterrows():
                corrections.append(
                    f"'{row['Office Name']}' (Sub Office '{so_name}') was tagged "
                    f"'{row['Sub Division Name']}' but its Branch Offices say "
                    f"'{majority_div}' — corrected to '{majority_div}'."
                )
            df.loc[df["Sub Office Name"] == so_name, "Sub Division Name"] = majority_div
        else:
            corrections.append(
                f"⚠️ '{so_name}' appears genuinely in multiple Sub Divisions "
                f"({', '.join(divisions)}) with disagreeing Branch Offices — "
                f"treated as distinct groups, NOT merged."
            )

    return df, corrections


# ── Main builder ─────────────────────────────────────────────────────────────
def build_report(master_file, consolidated_csv, single_date_csv,
                  consolidated_label: str, single_date_label: str,
                  file_tag: str) -> tuple[bytes, dict, list]:
    """
    Returns: (xlsx_bytes, division_totals_dict, corrections_list)
    division_totals_dict = {sub_division: {"cons_total":, "cons_digital":, "single_total":, "single_digital":}, ...}
                            plus a "KARAD DIVISION TOTAL" key with the grand totals.
    """
    master = load_master(master_file)
    master, corrections = reconcile_sub_division(master)

    cons_agg = load_cod_csv(consolidated_csv).set_index("Office ID")
    single_agg = load_cod_csv(single_date_csv).set_index("Office ID")

    master["C_Total"] = master["Office ID"].map(cons_agg["Total"])
    master["C_Digital"] = master["Office ID"].map(cons_agg["Digital"])
    master["S_Total"] = master["Office ID"].map(single_agg["Total"])
    master["S_Digital"] = master["Office ID"].map(single_agg["Digital"])

    master["Office Type"] = master["Office Type Code"].map(TYPE_DISPLAY)
    master["_type_rank"] = master["Office Type Code"].map(TYPE_SORT_RANK).fillna(9)
    master = master.sort_values(
        ["Sub Division Name", "Sub Office Name", "_type_rank", "Office Name"]
    ).reset_index(drop=True)
    master = master.drop(columns=["_type_rank"])

    sub_divisions = sorted(master["Sub Division Name"].dropna().unique().tolist())

    wb = Workbook()
    wb.remove(wb.active)

    # ============================================================
    # SHEET 1: Raw Data
    # ============================================================
    ws1 = wb.create_sheet("Raw Data")
    ws1.sheet_properties.tabColor = NAVY
    ws1.sheet_view.showGridLines = False

    row = 1
    _merge_write(ws1, row, 1, 11,
                 "KARAD DIVISION — COD DIGITAL TRANSACTION % REPORT — RAW DATA (OFFICE-WISE)",
                 _font(bold=True, size=TITLE_SIZE), _fill(WHITE), _align("center"), height=51)
    row += 1

    _merge_write(ws1, row, 1, 5, "", _font(size=HEADER_SIZE), _fill(WHITE), _align("center"))
    _merge_write(ws1, row, 6, 8, consolidated_label, _font(bold=True, color="FFFFFF", size=HEADER_SIZE),
                 _fill(DARK_GREEN), _align("center"))
    _merge_write(ws1, row, 9, 11, single_date_label, _font(bold=True, color="FFFFFF", size=HEADER_SIZE),
                 _fill(DARK_ORANGE), _align("center"))
    ws1.row_dimensions[row].height = 18
    row += 1

    headers1 = ["Sr. No.", "Sub Division Name", "Sub Office Name", "Office Name", "Office Type",
                "Total COD\nArticles", "COD Digital\nArticles", "Digital Txn %",
                "Total COD\nArticles", "COD Digital\nArticles", "Digital Txn %"]
    widths1 = [8, 18, 22, 22, 11, 14, 14, 12, 14, 14, 12]
    _write_header_row(ws1, row, headers1, widths1)
    row += 1

    raw_start = row
    for sr, (_, off) in enumerate(master.iterrows(), start=1):
        bg = WHITE if sr % 2 == 1 else LIGHT_BLUE
        c_total, c_dig = off["C_Total"], off["C_Digital"]
        s_total, s_dig = off["S_Total"], off["S_Digital"]

        c = ws1.cell(row=row, column=1, value=sr)
        c.font = _font(); c.fill = _fill(bg); c.alignment = _align("center"); c.border = ALL_BORDER

        for ci, val in enumerate(
            [off["Sub Division Name"], off["Sub Office Name"], off["Office Name"], off["Office Type"]], start=2
        ):
            halign = "left" if ci in (2, 3, 4) else "center"
            cell = ws1.cell(row=row, column=ci, value=val)
            cell.font = _font(); cell.fill = _fill(bg); cell.alignment = _align(halign); cell.border = ALL_BORDER

        f_cell = ws1.cell(row=row, column=6); f_cell.fill = _fill(bg); f_cell.alignment = _align("center"); f_cell.border = ALL_BORDER
        g_cell = ws1.cell(row=row, column=7); g_cell.fill = _fill(bg); g_cell.alignment = _align("center"); g_cell.border = ALL_BORDER
        h_cell = ws1.cell(row=row, column=8); h_cell.fill = _fill(bg); h_cell.alignment = _align("center"); h_cell.border = ALL_BORDER

        if pd.notna(c_total):
            f_cell.value = int(c_total); f_cell.font = _font(); f_cell.number_format = "#,##0"
            g_cell.value = int(c_dig); g_cell.font = _font(); g_cell.number_format = "#,##0"
            h_cell.value = f"=IF(F{row}=0,\"\",G{row}/F{row})"
            h_cell.number_format = "0.00%"
            is_zero_red = int(c_total) > 0 and int(c_dig) == 0
            h_cell.font = _font(bold=is_zero_red, color=RED if is_zero_red else "000000")
        else:
            f_cell.font = _font(); g_cell.font = _font(); h_cell.font = _font()

        i_cell = ws1.cell(row=row, column=9); i_cell.fill = _fill(bg); i_cell.alignment = _align("center"); i_cell.border = ALL_BORDER
        j_cell = ws1.cell(row=row, column=10); j_cell.fill = _fill(bg); j_cell.alignment = _align("center"); j_cell.border = ALL_BORDER
        k_cell = ws1.cell(row=row, column=11); k_cell.fill = _fill(bg); k_cell.alignment = _align("center"); k_cell.border = ALL_BORDER

        if pd.notna(s_total):
            i_cell.value = int(s_total); i_cell.font = _font(); i_cell.number_format = "#,##0"
            j_cell.value = int(s_dig); j_cell.font = _font(); j_cell.number_format = "#,##0"
            k_cell.value = f"=IF(I{row}=0,\"\",J{row}/I{row})"
            k_cell.number_format = "0.00%"
            is_zero_red_s = int(s_total) > 0 and int(s_dig) == 0
            k_cell.font = _font(bold=is_zero_red_s, color=RED if is_zero_red_s else "000000")
        else:
            i_cell.font = _font(); j_cell.font = _font(); k_cell.font = _font()

        ws1.row_dimensions[row].height = 20
        row += 1

    raw_end = row - 1
    # NOTE: no freeze_panes call — intentionally left unfrozen on every sheet.

    # ============================================================
    # SHEET 2: Sub Division wise Summary
    # ============================================================
    RAW = "'Raw Data'"
    ws2 = wb.create_sheet("Sub Division wise Summary")
    ws2.sheet_properties.tabColor = DARK_GREEN
    ws2.sheet_view.showGridLines = False

    row = 1
    _merge_write(ws2, row, 1, 9,
                 "KARAD DIVISION — SUB OFFICE & SUB DIVISION WISE COD DIGITAL TRANSACTION % SUMMARY",
                 _font(bold=True, size=TITLE_SIZE), _fill(WHITE), _align("center"), height=51)
    row += 1

    _merge_write(ws2, row, 1, 9, "TABLE A — SUB OFFICE WISE SUMMARY  (Sub Office + all Branch Offices under it)",
                 _font(bold=True, color="FFFFFF", size=SECTION_SIZE), _fill(NAVY), _align("left"), height=24)
    row += 1

    _merge_write(ws2, row, 1, 3, "", _font(size=HEADER_SIZE), _fill(WHITE), _align("center"))
    _merge_write(ws2, row, 4, 6, consolidated_label, _font(bold=True, color="FFFFFF", size=HEADER_SIZE),
                 _fill(DARK_GREEN), _align("center"))
    _merge_write(ws2, row, 7, 9, single_date_label, _font(bold=True, color="FFFFFF", size=HEADER_SIZE),
                 _fill(DARK_ORANGE), _align("center"))
    ws2.row_dimensions[row].height = 18
    row += 1

    headersA = ["Sr. No.", "Sub Division Name", "Sub Office Name",
                "Total COD\nArticles", "COD Digital\nArticles", "Digital Txn %",
                "Total COD\nArticles", "COD Digital\nArticles", "Digital Txn %"]
    widthsA = [8, 18, 24, 14, 14, 12, 14, 14, 12]
    _write_header_row(ws2, row, headersA, widthsA)
    row += 1

    sub_offices = master[["Sub Division Name", "Sub Office Name"]].drop_duplicates().reset_index(drop=True)
    tableA_start = row

    for i, so in sub_offices.iterrows():
        r = tableA_start + i
        bg = WHITE if i % 2 == 0 else LIGHT_BLUE
        subdiv, soname = so["Sub Division Name"], so["Sub Office Name"]

        c1 = ws2.cell(row=r, column=1, value=i + 1)
        c1.font = _font(); c1.fill = _fill(bg); c1.alignment = _align("center"); c1.border = ALL_BORDER
        c2 = ws2.cell(row=r, column=2, value=subdiv)
        c2.font = _font(); c2.fill = _fill(bg); c2.alignment = _align("left"); c2.border = ALL_BORDER
        c3 = ws2.cell(row=r, column=3, value=soname)
        c3.font = _font(); c3.fill = _fill(bg); c3.alignment = _align("left"); c3.border = ALL_BORDER

        # Dual-key SUMIFS — Sub Division + Sub Office Name together, never Sub Office Name alone
        d_cell = ws2.cell(row=r, column=4, value=f"=SUMIFS({RAW}!$F${raw_start}:$F${raw_end},{RAW}!$B${raw_start}:$B${raw_end},$B{r},{RAW}!$C${raw_start}:$C${raw_end},$C{r})")
        e_cell = ws2.cell(row=r, column=5, value=f"=SUMIFS({RAW}!$G${raw_start}:$G${raw_end},{RAW}!$B${raw_start}:$B${raw_end},$B{r},{RAW}!$C${raw_start}:$C${raw_end},$C{r})")
        f_cell = ws2.cell(row=r, column=6, value=f'=IF(D{r}=0,"",E{r}/D{r})')
        g_cell = ws2.cell(row=r, column=7, value=f"=SUMIFS({RAW}!$I${raw_start}:$I${raw_end},{RAW}!$B${raw_start}:$B${raw_end},$B{r},{RAW}!$C${raw_start}:$C${raw_end},$C{r})")
        h_cell = ws2.cell(row=r, column=8, value=f"=SUMIFS({RAW}!$J${raw_start}:$J${raw_end},{RAW}!$B${raw_start}:$B${raw_end},$B{r},{RAW}!$C${raw_start}:$C${raw_end},$C{r})")
        i_cell = ws2.cell(row=r, column=9, value=f'=IF(G{r}=0,"",H{r}/G{r})')

        for cell, fmt in [(d_cell, "#,##0"), (e_cell, "#,##0"), (g_cell, "#,##0"), (h_cell, "#,##0")]:
            cell.font = _font(); cell.fill = _fill(bg); cell.alignment = _align("center"); cell.border = ALL_BORDER
            cell.number_format = fmt
        for cell in (f_cell, i_cell):
            cell.font = _font(); cell.fill = _fill(bg); cell.alignment = _align("center"); cell.border = ALL_BORDER
            cell.number_format = "0.00%"

        ws2.row_dimensions[r].height = 20

    tableA_end = tableA_start + len(sub_offices) - 1
    row = tableA_end + 2

    _merge_write(ws2, row, 1, 8, "TABLE B — SUB DIVISION WISE SUMMARY (Volume-Weighted)",
                 _font(bold=True, color="FFFFFF", size=SECTION_SIZE), _fill(NAVY), _align("left"), height=24)
    row += 1

    _merge_write(ws2, row, 1, 2, "", _font(size=HEADER_SIZE), _fill(WHITE), _align("center"))
    _merge_write(ws2, row, 3, 5, consolidated_label, _font(bold=True, color="FFFFFF", size=HEADER_SIZE),
                 _fill(DARK_GREEN), _align("center"))
    _merge_write(ws2, row, 6, 8, single_date_label, _font(bold=True, color="FFFFFF", size=HEADER_SIZE),
                 _fill(DARK_ORANGE), _align("center"))
    ws2.row_dimensions[row].height = 18
    row += 1

    headersB = ["Sr. No.", "Sub Division Name",
                "Total COD\nArticles", "COD Digital\nArticles", "Digital Txn %",
                "Total COD\nArticles", "COD Digital\nArticles", "Digital Txn %"]
    widthsB = [8, 18, 14, 14, 12, 14, 14, 12]
    _write_header_row(ws2, row, headersB, widthsB)
    row += 1

    dataB_start = row
    for i, sd in enumerate(sub_divisions):
        r = dataB_start + i
        bg = WHITE if i % 2 == 0 else LIGHT_BLUE

        c1 = ws2.cell(row=r, column=1, value=i + 1)
        c1.font = _font(); c1.fill = _fill(bg); c1.alignment = _align("center"); c1.border = ALL_BORDER
        c2 = ws2.cell(row=r, column=2, value=sd)
        c2.font = _font(); c2.fill = _fill(bg); c2.alignment = _align("left"); c2.border = ALL_BORDER

        c_cell = ws2.cell(row=r, column=3, value=f"=SUMIF($B${tableA_start}:$B${tableA_end},$B{r},$D${tableA_start}:$D${tableA_end})")
        d_cell = ws2.cell(row=r, column=4, value=f"=SUMIF($B${tableA_start}:$B${tableA_end},$B{r},$E${tableA_start}:$E${tableA_end})")
        e_cell = ws2.cell(row=r, column=5, value=f'=IF(C{r}=0,"",D{r}/C{r})')
        f_cell = ws2.cell(row=r, column=6, value=f"=SUMIF($B${tableA_start}:$B${tableA_end},$B{r},$G${tableA_start}:$G${tableA_end})")
        g_cell = ws2.cell(row=r, column=7, value=f"=SUMIF($B${tableA_start}:$B${tableA_end},$B{r},$H${tableA_start}:$H${tableA_end})")
        h_cell = ws2.cell(row=r, column=8, value=f'=IF(F{r}=0,"",G{r}/F{r})')

        for cell, fmt in [(c_cell, "#,##0"), (d_cell, "#,##0"), (f_cell, "#,##0"), (g_cell, "#,##0")]:
            cell.font = _font(); cell.fill = _fill(bg); cell.alignment = _align("center"); cell.border = ALL_BORDER
            cell.number_format = fmt
        for cell in (e_cell, h_cell):
            cell.font = _font(); cell.fill = _fill(bg); cell.alignment = _align("center"); cell.border = ALL_BORDER
            cell.number_format = "0.00%"

        ws2.row_dimensions[r].height = 20

    dataB_end = dataB_start + len(sub_divisions) - 1
    div_row = dataB_end + 1

    name_cell = ws2.cell(row=div_row, column=1, value="")
    name_cell.font = _font(bold=True, color="FFFFFF"); name_cell.fill = _fill(GOLD); name_cell.alignment = _align("center"); name_cell.border = ALL_BORDER
    name2_cell = ws2.cell(row=div_row, column=2, value="KARAD DIVISION TOTAL")
    name2_cell.font = _font(bold=True, color="FFFFFF"); name2_cell.fill = _fill(GOLD); name2_cell.alignment = _align("left"); name2_cell.border = ALL_BORDER

    tc = ws2.cell(row=div_row, column=3, value=f"=SUM(C{dataB_start}:C{dataB_end})")
    dc = ws2.cell(row=div_row, column=4, value=f"=SUM(D{dataB_start}:D{dataB_end})")
    ec = ws2.cell(row=div_row, column=5, value=f'=IF(C{div_row}=0,"",D{div_row}/C{div_row})')
    fc = ws2.cell(row=div_row, column=6, value=f"=SUM(F{dataB_start}:F{dataB_end})")
    gc = ws2.cell(row=div_row, column=7, value=f"=SUM(G{dataB_start}:G{dataB_end})")
    hc = ws2.cell(row=div_row, column=8, value=f'=IF(F{div_row}=0,"",G{div_row}/F{div_row})')

    for cell, fmt in [(tc, "#,##0"), (dc, "#,##0"), (fc, "#,##0"), (gc, "#,##0")]:
        cell.font = _font(bold=True, color="FFFFFF"); cell.fill = _fill(GOLD); cell.alignment = _align("center"); cell.border = ALL_BORDER
        cell.number_format = fmt
    for cell in (ec, hc):
        cell.font = _font(bold=True, color="FFFFFF"); cell.fill = _fill(GOLD); cell.alignment = _align("center"); cell.border = ALL_BORDER
        cell.number_format = "0.00%"
    ws2.row_dimensions[div_row].height = 24

    # ============================================================
    # SHEET 3: Summary
    # ============================================================
    SD = "'Sub Division wise Summary'"
    sd_row_map = {sd: dataB_start + i for i, sd in enumerate(sub_divisions)}

    ws3 = wb.create_sheet("Summary")
    ws3.sheet_properties.tabColor = GOLD
    ws3.sheet_view.showGridLines = False

    row = 1
    _merge_write(ws3, row, 1, 6, "KARAD DIVISION — COD DIGITAL PAYMENT PERFORMANCE SUMMARY",
                 _font(bold=True, size=TITLE_SIZE), _fill(WHITE), _align("center"), height=51)
    row += 1

    _merge_write(ws3, row, 1, 6,
                 "This sheet summarises how many Cash-on-Delivery (COD) articles were received in each Sub "
                 "Division of Karad Division, how many of those were paid for using a digital payment mode "
                 "(UPI/QR), and what percentage that represents.",
                 _font(italic=True, color="595959", size=OFFICE_ROW_SIZE), _fill(WHITE), _align("left"), height=34)
    row += 2

    headers_s = ["Sub Division", "COD Articles\nReceived", "Delivered via\nDigital Payment", "Digital Txn %", "Delivered via\nCash", "Remarks"]
    widths_s = [20, 16, 18, 12, 14, 30]

    def write_summary_section(row, label, fill_color, col_letter_total, col_letter_digital, col_letter_pct):
        _merge_write(ws3, row, 1, 6, label, _font(bold=True, color="FFFFFF", size=SECTION_SIZE),
                     _fill(fill_color), _align("left"), height=24)
        row += 1
        _write_header_row(ws3, row, headers_s, widths_s)
        row += 1
        data_start = row
        for i, sd in enumerate(sub_divisions):
            r = data_start + i
            bg = WHITE if i % 2 == 0 else LIGHT_BLUE
            sd_src = sd_row_map[sd]

            nc = ws3.cell(row=r, column=1, value=sd)
            nc.font = _font(); nc.fill = _fill(bg); nc.alignment = _align("left"); nc.border = ALL_BORDER
            bc = ws3.cell(row=r, column=2, value=f"={SD}!{col_letter_total}{sd_src}")
            cc = ws3.cell(row=r, column=3, value=f"={SD}!{col_letter_digital}{sd_src}")
            dc = ws3.cell(row=r, column=4, value=f"={SD}!{col_letter_pct}{sd_src}")
            ec = ws3.cell(row=r, column=5, value=f"=B{r}-C{r}")
            fc = ws3.cell(row=r, column=6, value=(
                f'=IF(D{r}="","No COD activity recorded",'
                f'IF(D{r}>=0.5,"Healthy digital adoption",'
                f'IF(D{r}>=0.2,"Moderate — needs push","Low digital adoption — needs attention")))'
            ))
            for cell, fmt in [(bc, "#,##0"), (cc, "#,##0"), (ec, "#,##0")]:
                cell.font = _font(); cell.fill = _fill(bg); cell.alignment = _align("center"); cell.border = ALL_BORDER
                cell.number_format = fmt
            dc.font = _font(); dc.fill = _fill(bg); dc.alignment = _align("center"); dc.border = ALL_BORDER
            dc.number_format = "0.00%"
            fc.font = _font(); fc.fill = _fill(bg); fc.alignment = _align("left"); fc.border = ALL_BORDER
            ws3.row_dimensions[r].height = 22

        data_end = data_start + len(sub_divisions) - 1
        total_row = data_end + 1

        nc = ws3.cell(row=total_row, column=1, value="KARAD DIVISION TOTAL")
        nc.font = _font(bold=True, color="FFFFFF"); nc.fill = _fill(GOLD); nc.alignment = _align("left"); nc.border = ALL_BORDER
        bc = ws3.cell(row=total_row, column=2, value=f"={SD}!{col_letter_total}{div_row}")
        cc = ws3.cell(row=total_row, column=3, value=f"={SD}!{col_letter_digital}{div_row}")
        dc = ws3.cell(row=total_row, column=4, value=f"={SD}!{col_letter_pct}{div_row}")
        ec = ws3.cell(row=total_row, column=5, value=f"=B{total_row}-C{total_row}")
        fc = ws3.cell(row=total_row, column=6, value="Division-wide performance")
        for cell, fmt in [(bc, "#,##0"), (cc, "#,##0"), (ec, "#,##0")]:
            cell.font = _font(bold=True, color="FFFFFF"); cell.fill = _fill(GOLD); cell.alignment = _align("center"); cell.border = ALL_BORDER
            cell.number_format = fmt
        dc.font = _font(bold=True, color="FFFFFF"); dc.fill = _fill(GOLD); dc.alignment = _align("center"); dc.border = ALL_BORDER
        dc.number_format = "0.00%"
        fc.font = _font(bold=True, color="FFFFFF"); fc.fill = _fill(GOLD); fc.alignment = _align("left"); fc.border = ALL_BORDER
        ws3.row_dimensions[total_row].height = 24

        return total_row + 2

    row = write_summary_section(row, f"PERIOD: {consolidated_label}", DARK_GREEN, "C", "D", "E")
    row = write_summary_section(row, f"PERIOD: {single_date_label} (Single Day)", DARK_ORANGE, "F", "G", "H")

    _merge_write(ws3, row, 1, 6,
                 'Note: "Digital Txn %" = (COD articles paid via digital mode ÷ Total COD articles) × 100. '
                 "Sub Divisions with no COD activity in a given period show blank cells rather than 0%, since "
                 "there is no transaction base to calculate a percentage from.",
                 _font(italic=True, color="595959", size=OFFICE_ROW_SIZE), _fill(WHITE), _align("left"), height=34)

    # ── No freeze_panes set anywhere — intentional ──────────────────────────

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Build a lightweight totals dict for the Streamlit summary metrics
    cons_total = master["C_Total"].sum(skipna=True)
    cons_digital = master["C_Digital"].sum(skipna=True)
    single_total = master["S_Total"].sum(skipna=True)
    single_digital = master["S_Digital"].sum(skipna=True)

    totals = {
        "cons_total": int(cons_total),
        "cons_digital": int(cons_digital),
        "cons_pct": (cons_digital / cons_total * 100) if cons_total else None,
        "single_total": int(single_total),
        "single_digital": int(single_digital),
        "single_pct": (single_digital / single_total * 100) if single_total else None,
    }

    return output.read(), totals, corrections
